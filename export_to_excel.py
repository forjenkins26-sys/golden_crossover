"""
Export Delta Exchange Trades to Excel (Desktop)
"""

import pandas as pd
import os
from pathlib import Path
import warnings

warnings.filterwarnings('ignore')

print("\n[Creating Excel file...]")

# Read the CSV
csv_file = 'delta_exchange_trades.csv'
if not os.path.exists(csv_file):
    print(f"ERROR: {csv_file} not found. Run delta_exchange_fees.py first.")
    exit(1)

trades_df = pd.read_csv(csv_file)

# Get Desktop path
desktop = Path.home() / "Desktop"
excel_file = desktop / "RSI_30_70_Trades_Jan_Apr_2026.xlsx"

print(f"[Reading trades from {csv_file}...]")
print(f"[Total trades: {len(trades_df)}]")

# Create Excel writer with formatting
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    trades_df.to_excel(writer, sheet_name='Trades', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Trades']
    
    # Import styling libraries
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header
    for col_num, header in enumerate(trades_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Format data rows
    for row_num, row_data in enumerate(trades_df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color code by P&L
            if col_num == trades_df.columns.get_loc('Net_PnL') + 1:
                if isinstance(value, (int, float)) and value > 0:
                    cell.fill = win_fill
                    cell.font = Font(bold=True, color="006100")
                elif isinstance(value, (int, float)) and value < 0:
                    cell.fill = loss_fill
                    cell.font = Font(bold=True, color="9C0006")
            
            # Format numbers
            if col_num in [trades_df.columns.get_loc(col) + 1 for col in ['Entry_Price', 'Exit_Price', 'Gross_PnL', 'Entry_Cost', 'Exit_Cost', 'Total_Cost', 'Net_PnL', 'Cumulative_PnL']]:
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
            
            if col_num == trades_df.columns.get_loc('Return_%') + 1:
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
    
    # Adjust column widths
    column_widths = {
        'Date': 20,
        'Type': 8,
        'Entry_Price': 14,
        'Exit_Price': 14,
        'Result': 8,
        'Entry_RSI': 11,
        'Exit_RSI': 10,
        'Gross_PnL': 12,
        'Entry_Cost': 12,
        'Exit_Cost': 11,
        'Total_Cost': 12,
        'Net_PnL': 12,
        'Return_%': 10,
        'Duration_hrs': 14,
        'Cumulative_PnL': 15
    }
    
    for col_num, header in enumerate(trades_df.columns, 1):
        col_letter = get_column_letter(col_num)
        width = column_widths.get(header, 12)
        worksheet.column_dimensions[col_letter].width = width
    
    # Freeze top row
    worksheet.freeze_panes = 'A2'
    
    # Add summary sheet
    summary_sheet = workbook.create_sheet('Summary', 0)
    
    summary_sheet['A1'] = 'RSI 30/70 STRATEGY - SUMMARY'
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:B1')
    
    row = 3
    summary_data = [
        ['PERIOD', 'Jan 1 - Apr 10, 2026'],
        ['POSITION SIZE', '0.1 BTC per trade'],
        ['', ''],
        ['TOTAL TRADES', len(trades_df)],
        ['WINNING TRADES', len(trades_df[trades_df['Net_PnL'] > 0])],
        ['LOSING TRADES', len(trades_df[trades_df['Net_PnL'] < 0])],
        ['WIN RATE', f"{len(trades_df[trades_df['Net_PnL'] > 0])/len(trades_df)*100:.1f}%"],
        ['', ''],
        ['GROSS P&L', f"${trades_df['Gross_PnL'].sum():.2f}"],
        ['TOTAL FEES', f"${trades_df['Total_Cost'].sum():.2f}"],
        ['NET P&L', f"${trades_df['Net_PnL'].sum():.2f}"],
        ['', ''],
        ['AVERAGE WIN', f"${trades_df[trades_df['Net_PnL'] > 0]['Net_PnL'].mean():.2f}"],
        ['AVERAGE LOSS', f"${trades_df[trades_df['Net_PnL'] < 0]['Net_PnL'].mean():.2f}"],
        ['PROFIT FACTOR', f"{(trades_df[trades_df['Net_PnL'] > 0]['Net_PnL'].sum() / abs(trades_df[trades_df['Net_PnL'] < 0]['Net_PnL'].sum())):.2f}x"],
        ['BEST TRADE', f"${trades_df['Net_PnL'].max():.2f}"],
        ['WORST TRADE', f"${trades_df['Net_PnL'].min():.2f}"],
        ['', ''],
        ['LONG TRADES', len(trades_df[trades_df['Type'] == 'LONG'])],
        ['LONG WIN RATE', f"{len(trades_df[(trades_df['Type'] == 'LONG') & (trades_df['Net_PnL'] > 0)]) / len(trades_df[trades_df['Type'] == 'LONG']) * 100:.1f}%"],
        ['LONG P&L', f"${trades_df[trades_df['Type'] == 'LONG']['Net_PnL'].sum():.2f}"],
        ['', ''],
        ['SHORT TRADES', len(trades_df[trades_df['Type'] == 'SHORT'])],
        ['SHORT WIN RATE', f"{len(trades_df[(trades_df['Type'] == 'SHORT') & (trades_df['Net_PnL'] > 0)]) / len(trades_df[trades_df['Type'] == 'SHORT']) * 100:.1f}%"],
        ['SHORT P&L', f"${trades_df[trades_df['Type'] == 'SHORT']['Net_PnL'].sum():.2f}"],
    ]
    
    for item in summary_data:
        summary_sheet[f'A{row}'] = item[0]
        summary_sheet[f'B{row}'] = item[1]
        
        if item[0] != '':
            summary_sheet[f'A{row}'].font = Font(bold=True)
            summary_sheet[f'A{row}'].border = border
            summary_sheet[f'B{row}'].border = border
            summary_sheet[f'B{row}'].alignment = Alignment(horizontal='right')
        
        row += 1
    
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 25

print(f"[Excel file created successfully!]")
print(f"[Location: {excel_file}]")
print(f"[Total trades exported: {len(trades_df)}]")
print(f"\n✓ File saved to Desktop: RSI_30_70_Trades_Jan_Apr_2026.xlsx\n")
