# -*- coding: utf-8 -*-
"""
EXCEL_LOGGER.PY - Local Excel Journal for RSI Hybrid Bot
Handles daily Excel logging with automatic file creation and master aggregation
"""

import os
import json
from datetime import datetime, date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Base directory for all Excel logs
JOURNAL_BASE_DIR = Path.home() / "Desktop" / "RSI_Hybrid_Journal"
DAILY_LOGS_DIR = JOURNAL_BASE_DIR / "Daily_Logs"
MASTER_JOURNAL_PATH = JOURNAL_BASE_DIR / "Master_Journal.xlsx"

# Column headers for daily log
DAILY_LOG_HEADERS = [
    "Date", "Time", "Direction", "Entry Price", "Exit Price", 
    "Result", "Gross PnL", "Fees", "Net PnL", "Cumulative PnL", "Notes"
]

# Ensure directories exist
JOURNAL_BASE_DIR.mkdir(parents=True, exist_ok=True)
DAILY_LOGS_DIR.mkdir(parents=True, exist_ok=True)

def get_today_date_str():
    """Return today's date as YYYY-MM-DD"""
    return date.today().strftime("%Y-%m-%d")

def get_daily_log_path(date_str=None):
    """Get path to daily log file"""
    if date_str is None:
        date_str = get_today_date_str()
    return DAILY_LOGS_DIR / f"{date_str}.xlsx"

def create_daily_log_file(date_str=None):
    """Create a new daily Excel file if it doesn't exist"""
    if date_str is None:
        date_str = get_today_date_str()
    
    file_path = get_daily_log_path(date_str)
    
    if file_path.exists():
        return file_path
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trades"
    
    # Add headers
    for col_num, header in enumerate(DAILY_LOG_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = [12, 10, 10, 12, 12, 10, 12, 10, 12, 14, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add "No Trade Today" row if needed (will be replaced on first trade)
    ws.cell(row=2, column=1).value = date_str
    ws.cell(row=2, column=11).value = "No Trade Today"
    
    # Save workbook
    wb.save(file_path)
    print(f"[EXCEL] Created daily log: {file_path}")
    return file_path

def log_trade_to_excel(trade_data):
    """
    Log a trade to the daily Excel file
    trade_data dict should contain:
    {
        'date': 'YYYY-MM-DD',
        'time': 'HH:MM:SS',
        'direction': 'LONG' or 'SHORT',
        'entry_price': float,
        'exit_price': float,
        'result': 'WIN' or 'LOSS',
        'gross_pnl': float,
        'fees': float,
        'net_pnl': float,
        'cumulative_pnl': float,
        'notes': str
    }
    """
    date_str = trade_data.get('date', get_today_date_str())
    
    # Ensure daily file exists
    file_path = create_daily_log_file(date_str)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Find next empty row (skip header)
    next_row = ws.max_row + 1
    if ws.cell(row=2, column=11).value == "No Trade Today":
        next_row = 2  # Replace the "No Trade Today" row
        ws.delete_rows(2, 1)
        next_row = 2
    
    # Add trade data
    ws.cell(row=next_row, column=1).value = trade_data.get('date', date_str)
    ws.cell(row=next_row, column=2).value = trade_data.get('time', '')
    ws.cell(row=next_row, column=3).value = trade_data.get('direction', '')
    ws.cell(row=next_row, column=4).value = trade_data.get('entry_price', 0)
    ws.cell(row=next_row, column=5).value = trade_data.get('exit_price', 0)
    ws.cell(row=next_row, column=6).value = trade_data.get('result', '')
    ws.cell(row=next_row, column=7).value = trade_data.get('gross_pnl', 0)
    ws.cell(row=next_row, column=8).value = trade_data.get('fees', 0)
    ws.cell(row=next_row, column=9).value = trade_data.get('net_pnl', 0)
    ws.cell(row=next_row, column=10).value = trade_data.get('cumulative_pnl', 0)
    ws.cell(row=next_row, column=11).value = trade_data.get('notes', '')
    
    # Format numbers
    for col in [4, 5, 7, 8, 9, 10]:  # Price and PnL columns
        cell = ws.cell(row=next_row, column=col)
        cell.number_format = '0.00'
    
    wb.save(file_path)
    print(f"[EXCEL] Logged trade to {file_path}")

def create_master_journal():
    """Create or update Master_Journal.xlsx with aggregated data"""
    # Create new master workbook
    wb = openpyxl.Workbook()
    ws_trades = wb.active
    ws_trades.title = "All_Trades"
    ws_summary = wb.create_sheet("Summary")
    
    # Headers for All_Trades sheet
    for col_num, header in enumerate(DAILY_LOG_HEADERS, 1):
        cell = ws_trades.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = [12, 10, 10, 12, 12, 10, 12, 10, 12, 14, 20]
    for i, width in enumerate(column_widths, 1):
        ws_trades.column_dimensions[get_column_letter(i)].width = width
    
    # Aggregate data from daily files
    current_row = 2
    all_trades = []
    
    for daily_file in sorted(DAILY_LOGS_DIR.glob("*.xlsx")):
        daily_wb = openpyxl.load_workbook(daily_file)
        daily_ws = daily_wb.active
        
        for row in daily_ws.iter_rows(min_row=2, values_only=False):
            # Check if it's a trade (not "No Trade Today")
            notes_cell = row[10]  # Notes column
            if notes_cell.value != "No Trade Today":
                # Copy row to master
                for col_num, cell in enumerate(row, 1):
                    new_cell = ws_trades.cell(row=current_row, column=col_num)
                    new_cell.value = cell.value
                    new_cell.number_format = cell.number_format
                current_row += 1
                all_trades.append([cell.value for cell in row])
    
    # Create Summary sheet
    summary_headers = ["Metric", "Value"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Calculate metrics
    total_trades = len(all_trades)
    winning_trades = sum(1 for trade in all_trades if trade[5] == "WIN")
    losing_trades = total_trades - winning_trades
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    
    total_gross_pnl = sum(trade[6] if trade[6] else 0 for trade in all_trades)
    total_fees = sum(trade[7] if trade[7] else 0 for trade in all_trades)
    total_net_pnl = sum(trade[8] if trade[8] else 0 for trade in all_trades)
    
    avg_win = (sum(trade[6] for trade in all_trades if trade[5] == "WIN") / winning_trades) if winning_trades > 0 else 0
    avg_loss = abs(sum(trade[6] for trade in all_trades if trade[5] == "LOSS") / losing_trades) if losing_trades > 0 else 0
    profit_factor = (avg_win * winning_trades / (avg_loss * losing_trades)) if (avg_loss * losing_trades) > 0 else 0
    
    # Add metrics to summary sheet
    metrics = [
        ["Total Trades", total_trades],
        ["Winning Trades", winning_trades],
        ["Losing Trades", losing_trades],
        ["Win Rate (%)", round(win_rate, 2)],
        ["Gross PnL ($)", round(total_gross_pnl, 2)],
        ["Total Fees ($)", round(total_fees, 2)],
        ["Net PnL ($)", round(total_net_pnl, 2)],
        ["Avg Win ($)", round(avg_win, 2)],
        ["Avg Loss ($)", round(avg_loss, 2)],
        ["Profit Factor", round(profit_factor, 2)],
    ]
    
    for row_num, (metric, value) in enumerate(metrics, 2):
        ws_summary.cell(row=row_num, column=1).value = metric
        ws_summary.cell(row=row_num, column=2).value = value
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=2).number_format = '0.00'
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    # Save master journal
    wb.save(MASTER_JOURNAL_PATH)
    print(f"[EXCEL] Updated Master_Journal: {MASTER_JOURNAL_PATH}")

if __name__ == "__main__":
    # Test: Create today's daily log
    create_daily_log_file()
    print(f"Daily log path: {get_daily_log_path()}")
    print(f"Master journal path: {MASTER_JOURNAL_PATH}")
